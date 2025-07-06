import pandas as pd
import os
import sys
import shutil
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from urllib.parse import urlparse
from pathlib import Path
import tempfile

# Hide console window on Windows
if sys.platform == "win32":
    import subprocess
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

class TechAuditProcessor:
    def __init__(self):
        # Template file name
        self.template_name = "Template __ Tech Audit.xlsx"
        
        # Complete mapping of Item IDs to their data sources and calculations
        self.item_mappings = {
            # SITEMAP ISSUES
            "106": {"file": "internal_all.csv", "calculation": "urls_not_in_sitemap"},
            "107": {"file": "sitemap_all.csv", "calculation": "non_200_in_sitemap"},
            "108": {"file": "sitemap_all.csv", "calculation": "non_indexable_in_sitemap"},
            "109": {"file": "sitemap_all.csv", "calculation": "sitemap_timeout_errors"},
            "110": {"file": "sitemap_all.csv", "calculation": "large_sitemap_files"},
            
            # CANONICAL TAGS
            "51": {"file": "internal_all.csv", "calculation": "missing_canonical"},
            "52": {"file": "internal_all.csv", "calculation": "canonicalised_pages"},
            "53": {"file": "canonical_all.csv", "calculation": "non_indexable_canonical"},
            "54": {"file": "internal_all.csv", "calculation": "canonical_different_domain"},
            "55": {"file": "canonical_all.csv", "calculation": "missing_canonical_urls"},
            
            # CRAWLABILITY
            "56": {"file": "internal_all.csv", "calculation": "pages_with_noindex"},
            "57": {"file": "internal_all.csv", "calculation": "pages_with_nofollow"},
            "58": {"file": "internal_all.csv", "calculation": "conflicting_robots"},
            "59": {"file": "internal_all.csv", "calculation": "robots_txt_blocked"},
            
            # PAGE TITLES
            "1": {"file": "internal_all.csv", "calculation": "missing_page_titles"},
            "2": {"file": "internal_all.csv", "calculation": "duplicate_page_titles"},
            "3": {"file": "internal_all.csv", "calculation": "long_page_titles"},
            "4": {"file": "internal_all.csv", "calculation": "short_page_titles"},
            
            # META DESCRIPTIONS
            "7": {"file": "internal_all.csv", "calculation": "missing_meta_descriptions"},
            "8": {"file": "internal_all.csv", "calculation": "duplicate_meta_descriptions"},
            "9": {"file": "internal_all.csv", "calculation": "long_meta_descriptions"},
            "10": {"file": "internal_all.csv", "calculation": "short_meta_descriptions"},
            
            # H1 TAGS
            "13": {"file": "internal_all.csv", "calculation": "missing_h1"},
            "14": {"file": "internal_all.csv", "calculation": "duplicate_h1"},
            "15": {"file": "internal_all.csv", "calculation": "multiple_h1"},
            
            # IMAGES
            "70": {"file": "images_all.csv", "calculation": "images_missing_alt"},
            "72": {"file": "images_all.csv", "calculation": "images_over_100kb"},
            "130": {"file": "images_all.csv", "calculation": "broken_images"},
            
            # RESPONSE CODES
            "63": {"file": "internal_all.csv", "calculation": "client_4xx_errors"},
            "64": {"file": "internal_all.csv", "calculation": "server_5xx_errors"},
            "65": {"file": "internal_all.csv", "calculation": "status_404_count"},
            
            # REDIRECTS
            "18": {"file": "redirect_chains_all.csv", "calculation": "redirect_chains"},
            "19": {"file": "redirect_loops_all.csv", "calculation": "redirect_loops"},
            "20": {"file": "internal_all.csv", "calculation": "temporary_redirects"},
            
            # Add more mappings as needed
        }
        
        self.screaming_frog_data = {}
    
    def get_desktop_path(self):
        """Get the desktop path in a more reliable way"""
        # Try multiple methods to find Desktop
        possible_paths = [
            os.path.join(os.path.expanduser("~"), "Desktop"),
            os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),
            os.path.join(os.environ.get("USERPROFILE", ""), "Desktop"),
            os.path.join(os.environ.get("USERPROFILE", ""), "OneDrive", "Desktop"),
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                return path
        
        # If no desktop found, create a folder in user's home directory
        home = os.path.expanduser("~")
        audit_folder = os.path.join(home, "Tech Audit Reports")
        if not os.path.exists(audit_folder):
            os.makedirs(audit_folder)
        return audit_folder
        
    def process_audit(self, data_folder, client_name=""):
        """Main function to process the audit"""
        try:
            # Create timestamp for unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Create filename with client name
            if client_name:
                # Clean the client name (remove invalid filename characters)
                clean_client_name = "".join(c for c in client_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                output_filename = f"{clean_client_name}_Technical_Audit_{timestamp}.xlsx"
            else:
                output_filename = f"Technical_Audit_{timestamp}.xlsx"
            
            # Get output path
            output_folder = self.get_desktop_path()
            output_path = os.path.join(output_folder, output_filename)
            
            # Get template path
            template_path = self.get_template_path()
            
            print(f"Template found at: {template_path}")
            print(f"Output will be saved to: {output_path}")
            
            # Copy template to new file
            try:
                # Ensure output directory exists
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                shutil.copy2(template_path, output_path)
                print("Template copied successfully")
            except Exception as e:
                raise Exception(f"Failed to copy template to {output_path}: {str(e)}")
            
            # Load Screaming Frog data
            self.load_screaming_frog_data(data_folder)
            
            # Open the workbook and update values
            wb = load_workbook(output_path)
            
            # Update audit values
            print("Updating audit values...")
            self.update_audit_values(wb)
            
            # Import other Excel files from the folder
            print("Importing Excel files...")
            imported_count = self.import_existing_sheets(wb, data_folder)
            
            # Save the workbook
            print("Saving workbook...")
            wb.save(output_path)
            wb.close()
            
            # Return path and import count
            return output_path, imported_count
            
        except Exception as e:
            # If output file was created but error occurred, try to delete it
            if 'output_path' in locals() and os.path.exists(output_path):
                try:
                    os.remove(output_path)
                except:
                    pass
            raise e
    
    def get_template_path(self):
        """Get the full path to the template file - creates temp file from embedded data if needed"""
        # When running as compiled exe
        if getattr(sys, 'frozen', False):
            # First try to find the template in the PyInstaller bundle
            try:
                # This is where PyInstaller puts bundled files
                bundled_template = os.path.join(sys._MEIPASS, self.template_name)
                if os.path.exists(bundled_template):
                    return bundled_template
            except:
                pass
            
            # If not bundled, check next to the exe
            exe_dir = os.path.dirname(sys.executable)
            template_next_to_exe = os.path.join(exe_dir, self.template_name)
            if os.path.exists(template_next_to_exe):
                return template_next_to_exe
            
            # If still not found, extract from embedded data
            return self.create_template_from_embedded()
        
        else:
            # Running as script - check script directory
            script_dir = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(script_dir, self.template_name)
            if os.path.exists(template_path):
                return template_path
            
            # If not found, create from embedded data
            return self.create_template_from_embedded()
    
    def create_template_from_embedded(self):
        """Create the template file from embedded data if it doesn't exist"""
        # Create a temporary file for the template
        temp_dir = tempfile.gettempdir()
        temp_template_path = os.path.join(temp_dir, self.template_name)
        
        # If temp template already exists and is recent (less than 1 hour old), use it
        if os.path.exists(temp_template_path):
            file_age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(temp_template_path))
            if file_age.total_seconds() < 3600:  # Less than 1 hour old
                return temp_template_path
        
        # Create a basic template structure
        # This is a fallback - ideally the template should be bundled with PyInstaller
        wb = openpyxl.Workbook()
        
        # Create Full Audit sheet
        ws = wb.active
        ws.title = "Full Audit"
        
        # Add headers
        headers = ["Sort", "Checked", "Item ID", "Issue Name", "Column 19", 
                   "SF Error Name (For SF Issues)", "Parent Category", "Pass/Fail", 
                   "Expected Value", "Audit Value", "Priority"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Add some basic audit items
        audit_items = [
            ["1", "x", "1", "Missing Page Titles", "", "", "SEO", "", "0", "", ""],
            ["2", "x", "2", "Duplicate Page Titles", "", "", "SEO", "", "0", "", ""],
            ["3", "x", "3", "Long Page Titles", "", "", "SEO", "", "0", "", ""],
            ["4", "x", "4", "Short Page Titles", "", "", "SEO", "", "0", "", ""],
            ["7", "x", "7", "Missing Meta Descriptions", "", "", "SEO", "", "0", "", ""],
            ["8", "x", "8", "Duplicate Meta Descriptions", "", "", "SEO", "", "0", "", ""],
            ["13", "x", "13", "Missing H1", "", "", "SEO", "", "0", "", ""],
            ["14", "x", "14", "Duplicate H1", "", "", "SEO", "", "0", "", ""],
            ["15", "x", "15", "Multiple H1", "", "", "SEO", "", "0", "", ""],
            ["51", "x", "51", "Missing Canonical Tags", "", "", "Indexation", "", "0", "", ""],
            ["56", "x", "56", "Pages with Noindex", "", "", "Indexation", "", "0", "", ""],
            ["63", "x", "63", "4xx Errors", "", "", "Technical", "", "0", "", ""],
            ["64", "x", "64", "5xx Errors", "", "", "Technical", "", "0", "", ""],
            ["65", "x", "65", "404 Errors", "", "", "Technical", "", "0", "", ""],
        ]
        
        for row_num, item_data in enumerate(audit_items, 2):
            for col_num, value in enumerate(item_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Create Opportunities sheet
        wb.create_sheet("Opportunities")
        
        # Save the template
        wb.save(temp_template_path)
        wb.close()
        
        return temp_template_path
    
    def load_screaming_frog_data(self, data_folder):
        """Load all relevant Screaming Frog CSV files"""
        print("Loading Screaming Frog data...")
        
        # List of all possible files we might need
        files_to_load = [
            'internal_all.csv',
            'external_all.csv',
            'response_codes_all.csv',
            'page_titles_all.csv',
            'meta_descriptions_all.csv',
            'h1_all.csv',
            'h2_all.csv',
            'images_all.csv',
            'canonical_all.csv',
            'directives_all.csv',
            'structured_data_all.csv',
            'sitemap_all.csv',
            'redirect_chains_all.csv',
            'redirect_loops_all.csv'
        ]
        
        for file_name in files_to_load:
            file_path = os.path.join(data_folder, file_name)
            if os.path.exists(file_path):
                try:
                    self.screaming_frog_data[file_name] = pd.read_csv(file_path, low_memory=False)
                    print(f"  Loaded {file_name}: {len(self.screaming_frog_data[file_name])} rows")
                except Exception as e:
                    print(f"  Error loading {file_name}: {str(e)}")
            else:
                print(f"  {file_name} not found (optional)")
    
    def update_audit_values(self, wb):
        """Update the audit values in the workbook"""
        # Get the Full Audit sheet
        if 'Full Audit' not in wb.sheetnames:
            raise ValueError("'Full Audit' sheet not found in template")
        
        ws = wb['Full Audit']
        
        # Iterate through rows to find items to update
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
            item_id = ws.cell(row=row, column=3).value  # Column C is Item ID
            
            if item_id and str(item_id) in self.item_mappings:
                # Calculate the value for this item
                mapping = self.item_mappings[str(item_id)]
                value = self.calculate_metric(mapping['file'], mapping['calculation'])
                
                # Update the Audit Value (column J)
                ws.cell(row=row, column=10).value = value
                
                # Update Pass/Fail status (column H) based on Expected Value (column I)
                expected_value = ws.cell(row=row, column=9).value
                
                # Handle different expected value formats
                if expected_value is not None:
                    if str(expected_value).strip() == "0":
                        # Expected value is 0
                        if value == 0:
                            ws.cell(row=row, column=8).value = "Pass"
                            ws.cell(row=row, column=11).value = "N/A - Pass"  # Update Priority
                        else:
                            ws.cell(row=row, column=8).value = "Fail"
                            # Keep existing priority or set based on severity
                    elif str(expected_value).isdigit():
                        # Expected value is a number
                        expected = int(expected_value)
                        if value <= expected:
                            ws.cell(row=row, column=8).value = "Pass"
                            ws.cell(row=row, column=11).value = "N/A - Pass"
                        else:
                            ws.cell(row=row, column=8).value = "Fail"
                    elif "manual" in str(expected_value).lower():
                        # Manual review required
                        ws.cell(row=row, column=8).value = "Opportunity"
                        # Keep audit value for review
        
        print("Audit values updated successfully")
    
    def calculate_metric(self, file_name, calculation_type):
        """Calculate specific metrics from Screaming Frog data"""
        if file_name not in self.screaming_frog_data:
            return 0  # Return 0 if file not found
        
        df = self.screaming_frog_data[file_name]
        
        try:
            # SITEMAP CALCULATIONS
            if calculation_type == "urls_not_in_sitemap":
                # This requires comparing URLs - implement if both files exist
                return 0
            
            elif calculation_type == "non_200_in_sitemap":
                if 'Status Code' in df.columns:
                    return len(df[df['Status Code'] != 200])
                return 0
            
            elif calculation_type == "non_indexable_in_sitemap":
                if 'Indexability' in df.columns:
                    return len(df[df['Indexability'] != 'Indexable'])
                return 0
            
            # CANONICAL CALCULATIONS
            elif calculation_type == "missing_canonical":
                if 'Canonical Link Element 1' in df.columns:
                    # Only count HTML pages
                    if 'Content Type' in df.columns:
                        html_df = df[df['Content Type'].str.contains('text/html', na=False)]
                    else:
                        html_df = df
                    return len(html_df[html_df['Canonical Link Element 1'].isna()])
                return 0
            
            elif calculation_type == "canonicalised_pages":
                if 'Canonical Link Element 1' in df.columns and 'Address' in df.columns:
                    canonicalised = df[
                        (df['Canonical Link Element 1'].notna()) & 
                        (df['Canonical Link Element 1'] != df['Address'])
                    ]
                    return len(canonicalised)
                return 0
            
            elif calculation_type == "canonical_different_domain":
                if 'Canonical Link Element 1' in df.columns and 'Address' in df.columns:
                    def get_domain(url):
                        try:
                            return urlparse(str(url)).netloc
                        except:
                            return ''
                    
                    df_copy = df.copy()
                    df_copy['Page Domain'] = df_copy['Address'].apply(get_domain)
                    df_copy['Canonical Domain'] = df_copy['Canonical Link Element 1'].apply(get_domain)
                    
                    different_domain = df_copy[
                        (df_copy['Canonical Link Element 1'].notna()) &
                        (df_copy['Page Domain'] != df_copy['Canonical Domain']) &
                        (df_copy['Canonical Domain'] != '')
                    ]
                    return len(different_domain)
                return 0
            
            # CRAWLABILITY CALCULATIONS
            elif calculation_type == "pages_with_noindex":
                if 'Meta Robots 1' in df.columns:
                    noindex_pages = df[df['Meta Robots 1'].str.contains('noindex', na=False, case=False)]
                    return len(noindex_pages)
                return 0
            
            elif calculation_type == "pages_with_nofollow":
                if 'Meta Robots 1' in df.columns:
                    nofollow_pages = df[df['Meta Robots 1'].str.contains('nofollow', na=False, case=False)]
                    return len(nofollow_pages)
                return 0
            
            elif calculation_type == "robots_txt_blocked":
                if 'Indexability' in df.columns:
                    blocked = df[df['Indexability'].str.contains('Blocked by robots.txt', na=False, case=False)]
                    return len(blocked)
                return 0
            
            # PAGE TITLE CALCULATIONS
            elif calculation_type == "missing_page_titles":
                if 'Title 1' in df.columns:
                    missing = df[df['Title 1'].isna() | (df['Title 1'] == '')]
                    return len(missing)
                return 0
            
            elif calculation_type == "duplicate_page_titles":
                if 'Title 1' in df.columns:
                    # Only count non-empty titles
                    non_empty = df[df['Title 1'].notna() & (df['Title 1'] != '')]
                    duplicates = non_empty[non_empty.duplicated(subset=['Title 1'], keep=False)]
                    return len(duplicates)
                return 0
            
            elif calculation_type == "long_page_titles":
                if 'Title 1 Length' in df.columns:
                    long_titles = df[df['Title 1 Length'] > 60]
                    return len(long_titles)
                return 0
            
            elif calculation_type == "short_page_titles":
                if 'Title 1 Length' in df.columns:
                    short_titles = df[(df['Title 1 Length'] < 30) & (df['Title 1 Length'] > 0)]
                    return len(short_titles)
                return 0
            
            # META DESCRIPTION CALCULATIONS
            elif calculation_type == "missing_meta_descriptions":
                if 'Meta Description 1' in df.columns:
                    missing = df[df['Meta Description 1'].isna() | (df['Meta Description 1'] == '')]
                    return len(missing)
                return 0
            
            elif calculation_type == "duplicate_meta_descriptions":
                if 'Meta Description 1' in df.columns:
                    non_empty = df[df['Meta Description 1'].notna() & (df['Meta Description 1'] != '')]
                    duplicates = non_empty[non_empty.duplicated(subset=['Meta Description 1'], keep=False)]
                    return len(duplicates)
                return 0
            
            elif calculation_type == "long_meta_descriptions":
                if 'Meta Description 1 Length' in df.columns:
                    long_meta = df[df['Meta Description 1 Length'] > 160]
                    return len(long_meta)
                return 0
            
            elif calculation_type == "short_meta_descriptions":
                if 'Meta Description 1 Length' in df.columns:
                    short_meta = df[(df['Meta Description 1 Length'] < 120) & (df['Meta Description 1 Length'] > 0)]
                    return len(short_meta)
                return 0
            
            # H1 CALCULATIONS
            elif calculation_type == "missing_h1":
                if 'H1-1' in df.columns:
                    missing = df[df['H1-1'].isna() | (df['H1-1'] == '')]
                    return len(missing)
                return 0
            
            elif calculation_type == "duplicate_h1":
                if 'H1-1' in df.columns:
                    non_empty = df[df['H1-1'].notna() & (df['H1-1'] != '')]
                    duplicates = non_empty[non_empty.duplicated(subset=['H1-1'], keep=False)]
                    return len(duplicates)
                return 0
            
            elif calculation_type == "multiple_h1":
                if 'H1-2' in df.columns:
                    multiple = df[df['H1-2'].notna()]
                    return len(multiple)
                return 0
            
            # IMAGE CALCULATIONS
            elif calculation_type == "images_missing_alt":
                if 'Alt Text' in df.columns:
                    missing_alt = df[df['Alt Text'].isna() | (df['Alt Text'] == '')]
                    return len(missing_alt)
                return 0
            
            elif calculation_type == "images_over_100kb":
                if 'Size (Bytes)' in df.columns:
                    large_images = df[df['Size (Bytes)'] > 100000]
                    return len(large_images)
                return 0
            
            elif calculation_type == "broken_images":
                if 'Status Code' in df.columns:
                    broken = df[df['Status Code'] != 200]
                    return len(broken)
                return 0
            
            # RESPONSE CODE CALCULATIONS
            elif calculation_type == "client_4xx_errors":
                if 'Status Code' in df.columns:
                    errors_4xx = df[(df['Status Code'] >= 400) & (df['Status Code'] < 500)]
                    return len(errors_4xx)
                return 0
            
            elif calculation_type == "server_5xx_errors":
                if 'Status Code' in df.columns:
                    errors_5xx = df[df['Status Code'] >= 500]
                    return len(errors_5xx)
                return 0
            
            elif calculation_type == "status_404_count":
                if 'Status Code' in df.columns:
                    return len(df[df['Status Code'] == 404])
                return 0
            
            # REDIRECT CALCULATIONS
            elif calculation_type == "redirect_chains":
                # If we have the redirect chains file
                if 'redirect_chains_all.csv' in self.screaming_frog_data:
                    chains_df = self.screaming_frog_data['redirect_chains_all.csv']
                    return len(chains_df)
                return 0
            
            elif calculation_type == "redirect_loops":
                # If we have the redirect loops file
                if 'redirect_loops_all.csv' in self.screaming_frog_data:
                    loops_df = self.screaming_frog_data['redirect_loops_all.csv']
                    return len(loops_df)
                return 0
            
            elif calculation_type == "temporary_redirects":
                if 'Status Code' in df.columns:
                    temp_redirects = df[(df['Status Code'] == 302) | (df['Status Code'] == 307)]
                    return len(temp_redirects)
                return 0
            
            # Default return
            return 0
            
        except Exception as e:
            print(f"Error calculating {calculation_type}: {str(e)}")
            return 0
    
    def import_existing_sheets(self, workbook, folder_path):
        """Import all Excel files from the folder as new sheets"""
        print("Looking for Excel files to import...")
        
        # Get the name of the current output file to skip it
        current_file = os.path.basename(workbook.properties.title) if workbook.properties.title else None
        
        # Find all Excel files in the folder
        excel_files = []
        for file in os.listdir(folder_path):
            if file.lower().endswith(('.xlsx', '.xls')) and not file.startswith('~'):
                # Skip temporary Excel files (start with ~) and files that look like our output
                if not file.startswith(('Technical_Audit_', 'Tech_Audit_')) and 'Technical_Audit' not in file:
                    excel_files.append(file)
        
        if not excel_files:
            print("No Excel files found to import")
            return 0
        
        print(f"Found {len(excel_files)} Excel file(s) to import")
        
        # Get existing sheet names to avoid conflicts
        existing_sheets = set(workbook.sheetnames)
        imported_count = 0
        
        for excel_file in excel_files:
            try:
                file_path = os.path.join(folder_path, excel_file)
                print(f"\nImporting: {excel_file}")
                
                # Open the source workbook
                # Use data_only=True to get calculated values instead of formulas
                source_wb = load_workbook(file_path, data_only=True)
                
                # Copy each sheet from the source workbook
                for sheet_name in source_wb.sheetnames:
                    # Create unique sheet name if needed
                    new_sheet_name = sheet_name
                    counter = 1
                    while new_sheet_name in existing_sheets:
                        # Add number suffix if sheet name already exists
                        new_sheet_name = f"{sheet_name}_{counter}"
                        counter += 1
                    
                    # Create new sheet in target workbook
                    new_sheet = workbook.create_sheet(new_sheet_name)
                    existing_sheets.add(new_sheet_name)
                    
                    # Get source sheet
                    source_sheet = source_wb[sheet_name]
                    
                    # Copy merged cells if any
                    if source_sheet.merged_cells:
                        for merged_range in source_sheet.merged_cells.ranges:
                            new_sheet.merge_cells(str(merged_range))
                    
                    # Copy all cells from source to target
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                            
                            # Copy value
                            new_cell.value = cell.value
                            
                            # Copy basic formatting if available
                            try:
                                if cell.has_style:
                                    new_cell.font = openpyxl.styles.Font(
                                        bold=cell.font.bold if cell.font else False,
                                        italic=cell.font.italic if cell.font else False,
                                        color=cell.font.color if cell.font else None
                                    )
                                    if cell.fill and cell.fill.fill_type:
                                        new_cell.fill = openpyxl.styles.PatternFill(
                                            fill_type=cell.fill.fill_type,
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color
                                        )
                                    if cell.alignment:
                                        new_cell.alignment = openpyxl.styles.Alignment(
                                            horizontal=cell.alignment.horizontal,
                                            vertical=cell.alignment.vertical,
                                            wrap_text=cell.alignment.wrap_text
                                        )
                            except:
                                # Skip formatting if there's any error
                                pass
                    
                    # Copy column widths
                    for column in source_sheet.column_dimensions:
                        new_sheet.column_dimensions[column].width = source_sheet.column_dimensions[column].width
                    
                    # Copy row heights
                    for row in source_sheet.row_dimensions:
                        new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                    
                    print(f"  - Imported sheet '{sheet_name}' as '{new_sheet_name}'")
                
                source_wb.close()
                imported_count += 1
                
            except Exception as e:
                print(f"  - Error importing {excel_file}: {str(e)}")
                continue
        
        print(f"\nExcel file import complete - imported {imported_count} file(s)")
        return imported_count


class TechAuditGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Tech Audit Processor")
        self.root.geometry("600x470")
        self.root.configure(bg='#f0f0f0')
        
        # Title
        title_label = tk.Label(root, text="Tech Audit Processor", 
                              font=("Arial", 18, "bold"), bg='#f0f0f0')
        title_label.pack(pady=20)
        
        # Instructions
        instructions = tk.Label(root, 
                               text="This tool will analyze your Screaming Frog exports\n" +
                                    "and create a comprehensive technical audit report\n" +
                                    "Any Excel files in the folder will be imported as additional tabs",
                               font=("Arial", 11), bg='#f0f0f0', justify="center")
        instructions.pack(pady=10)
        
        # Template status
        template_frame = tk.Frame(root, bg='#f0f0f0')
        template_frame.pack(pady=5)
        
        template_label = tk.Label(template_frame, text="Template:", 
                                 font=("Arial", 10), bg='#f0f0f0')
        template_label.pack(side="left")
        
        self.template_status = tk.Label(template_frame, text="✓ Ready", 
                                       font=("Arial", 10, "bold"), bg='#f0f0f0', fg='#70AD47')
        self.template_status.pack(side="left", padx=(5, 0))
        
        # Client name input
        client_frame = tk.Frame(root, bg='#f0f0f0')
        client_frame.pack(pady=15, padx=50, fill="x")
        
        client_label = tk.Label(client_frame, text="Client Name:", 
                               font=("Arial", 10), bg='#f0f0f0')
        client_label.pack(side="left", padx=(0, 10))
        
        self.client_name_var = tk.StringVar()
        self.client_entry = tk.Entry(client_frame, textvariable=self.client_name_var, 
                                    font=("Arial", 10), width=30)
        self.client_entry.pack(side="left")
        
        # Folder selection
        folder_frame = tk.Frame(root, bg='#f0f0f0')
        folder_frame.pack(pady=15, padx=50, fill="x")
        
        folder_label = tk.Label(folder_frame, text="Data Folder:", 
                               font=("Arial", 10), bg='#f0f0f0')
        folder_label.pack(side="left", padx=(0, 10))
        
        self.folder_path_var = tk.StringVar()
        self.folder_entry = tk.Entry(folder_frame, textvariable=self.folder_path_var, 
                                    font=("Arial", 10), width=30)
        self.folder_entry.pack(side="left", padx=(0, 10))
        
        self.browse_button = tk.Button(folder_frame, text="Browse...", 
                                      command=self.browse_folder,
                                      bg='#4472C4', fg='white', 
                                      font=("Arial", 10, "bold"))
        self.browse_button.pack(side="left")
        
        # Process button
        self.process_button = tk.Button(root, text="Run Tech Audit", 
                                       command=self.process_audit,
                                       bg='#70AD47', fg='white', 
                                       font=("Arial", 14, "bold"),
                                       padx=30, pady=10)
        self.process_button.pack(pady=20)
        
        # Progress bar
        self.progress = ttk.Progressbar(root, length=400, mode='indeterminate')
        self.progress.pack(pady=10)
        
        # Status label
        self.status_label = tk.Label(root, text="Ready to process", 
                                    font=("Arial", 10), bg='#f0f0f0')
        self.status_label.pack(pady=10)
        
        # Output location note
        self.processor = TechAuditProcessor()
        output_location = self.processor.get_desktop_path()
        location_text = "Desktop" if "Desktop" in output_location else os.path.basename(output_location)
        
        note_label = tk.Label(root, 
                             text=f"Note: Reports will be saved to your {location_text}",
                             font=("Arial", 9, "italic"), bg='#f0f0f0', fg='#666666')
        note_label.pack(pady=5)
    
    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path_var.set(folder_selected)
            self.status_label.config(text="Folder selected")
    
    def process_audit(self):
        folder_path = self.folder_path_var.get()
        client_name = self.client_name_var.get()
        
        if not folder_path:
            messagebox.showerror("Error", "Please select a folder containing Screaming Frog exports!")
            return
        
        if not client_name:
            messagebox.showerror("Error", "Please enter a client name!")
            return
        
        # Disable buttons and inputs and start progress
        self.process_button.config(state="disabled")
        self.browse_button.config(state="disabled")
        self.client_entry.config(state="disabled")
        self.progress.start()
        self.status_label.config(text="Processing... Please wait")
        
        # Run in separate thread
        thread = threading.Thread(target=self.run_processor, args=(folder_path, client_name))
        thread.start()
    
    def run_processor(self, folder_path, client_name):
        try:
            processor = TechAuditProcessor()
            result = processor.process_audit(folder_path, client_name)
            
            # Handle both single value and tuple return
            if isinstance(result, tuple):
                output_file, imported_count = result
            else:
                output_file = result
                imported_count = 0
            
            # Update GUI in main thread
            self.root.after(0, self.processing_complete, True, (output_file, imported_count))
        except Exception as e:
            self.root.after(0, self.processing_complete, False, str(e))
    
    def processing_complete(self, success, message):
        self.progress.stop()
        self.process_button.config(state="normal")
        self.browse_button.config(state="normal")
        self.client_entry.config(state="normal")
        
        if success:
            self.status_label.config(text="Audit complete!")
            
            # Extract file path and import count
            if isinstance(message, tuple):
                output_file, imported_count = message
            else:
                output_file = message
                imported_count = 0
            
            output_location = os.path.dirname(output_file)
            location_name = "Desktop" if "Desktop" in output_location else os.path.basename(output_location)
            
            # Build success message
            success_msg = f"Tech audit completed successfully!\n\n"
            success_msg += f"Report saved to {location_name} as:\n{os.path.basename(output_file)}"
            
            if imported_count > 0:
                success_msg += f"\n\nImported {imported_count} Excel file(s) as additional tabs"
            
            messagebox.showinfo("Success", success_msg)
        else:
            self.status_label.config(text="Error occurred")
            messagebox.showerror("Error", f"An error occurred:\n\n{message}")


def main():
    root = tk.Tk()
    app = TechAuditGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()